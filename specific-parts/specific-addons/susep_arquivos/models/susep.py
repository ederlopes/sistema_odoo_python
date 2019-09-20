# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.addons.susep_arquivos.scripts.arquivos import SUSEP
from odoo.exceptions import ValidationError as UserError
from tempfile import TemporaryFile
import pandas as pd
import base64
import sys, os
from openpyxl import load_workbook
from time import gmtime, strftime
import dbf

CONVERT_DADOS = {"C": str, "N": int, "F": float, "D": str}


class Susep(models.Model):
    _name = 'susep.arquivos'
    _description = 'Arquivos susep'
    _rec_name = 'descricao'

    arquivo = fields.Binary(
        string='arquivo'
    )

    arquivo_txt = fields.Binary(
        string='arquivo txt',
        readonly=True
    )

    nome_txt = fields.Char(
        string='Arquivo TXT gerado',
        size=64
    )

    nome_arquivo = fields.Char(
        string="Nome do Arquivo"
    )

    descricao = fields.Char(
        string=u'Descrição',
    )

    data = fields.Datetime(
        string=u'Data de Cadastro',
    )

    id_user = fields.Many2one(
        'res.users',
        'Usuario',
        default=lambda self: self.env.user
    )

    quadro_id = fields.Many2one(
        comodel_name='susep.quadro',
        string='Quadro',
    )

    def _ler_arquivo(self, arquivo, sheet_name, param_converters=False):
        file = base64.b64decode(arquivo)
        fileobj = TemporaryFile('wb+')
        fileobj.write(file)
        fileobj.seek(0)

        wb = load_workbook(fileobj, read_only=True)

        if not sheet_name in wb.sheetnames:
            raise UserError(
                u'A planilha não contém o tab do quadro ' + sheet_name)

        if param_converters:
            return pd.read_excel(fileobj, sheet_name=sheet_name,
                                 converters=param_converters)
        else:
            return pd.read_excel(fileobj, sheet_name=sheet_name,
                                 converters={'ESPSEQ': str,
                                             'ENTCODIGO': str,
                                             'MRFMESANO': str,
                                             'QUAID': str,
                                             'TPMOID': str,
                                             'CMPID': str,
                                             'RAMCODIGO': str,
                                             'ESPDATAINICIORO': str,
                                             'ESPDATAFIMRO': str,
                                             'ESPDATAEMISSRO': str,
                                             'ESPVALORMOVRO': str,
                                             'ESPDATAINICIORD': str,
                                             'ESPDATAFIMRD': str,
                                             'ESPDATAEMISSRD': str,
                                             'ESPVALORMOVRD': str,
                                             'ESPCODCESS': str,
                                             'ESPFREQ': str,
                                             'ESPVARCARO': str,
                                             'ESPVALORCARD': str,
                                             'ESPVALORCIRO': str,
                                             'ESPVALORCIRD': str,
                                             'ESPMOEDA': str})

    def criar_quadro_txt_susep(self, res, arquivo):

        nome_arquivo_final = res.quadro_id.nome_arquivo_exportacao + '.txt'
        susep = SUSEP()
        resultado = ''
        i = 0
        arquivo = self._ler_arquivo(arquivo, res.quadro_id.tab_excel)

        arquivo.iterrows()

        for index, col in arquivo.iterrows():
            susep.ESPSEQ = col['ESPSEQ']
            # susep.ENTCODIGO = col['ENTCODIGO']
            susep.MRFMESANO = col['MRFMESANO']
            # susep.QUAID = col['QUAID']
            susep.TPMOID = col['TPMOID']
            susep.CMPID = col['CMPID']
            # susep.RAMCODIGO = col['RAMCODIGO']
            susep.ESPDATAINICIORO = col['ESPDATAINICIORO']
            susep.ESPDATAFIMRO = col['ESPDATAFIMRO']
            susep.ESPDATAEMISSRO = col['ESPDATAEMISSRO']
            susep.ESPVALORMOVRO = col['ESPVALORMOVRO']
            susep.ESPDATAINICIORD = col['ESPDATAINICIORD']
            susep.ESPDATAEMISSRD = col['ESPDATAEMISSRD']

            susep.ESPDATAFIMRD = col['ESPDATAFIMRD']
            susep.ESPDATAEMISSRD = col['ESPDATAEMISSRD']
            susep.ESPVALORMOVRD = col['ESPVALORMOVRD']
            # susep.ESPCODCESS = col['ESPCODCESS']
            susep.ESPFREQ = col['ESPFREQ']
            susep.ESPVARCARO = col['ESPVARCARO']
            susep.ESPVALORCARD = col['ESPVALORCARD']
            susep.ESPVALORCIRO = col['ESPVALORCIRO']
            susep.ESPVALORCIRD = col['ESPVALORCIRD']
            susep.ESPMOEDA = col['ESPMOEDA']

            resultado += susep._arquivo_quadro_378(res.quadro_id.numero_linhas)

        arquivo = open(nome_arquivo_final, 'w+')
        arquivo.write(resultado)
        arquivo.close()

        file = open(nome_arquivo_final, "rb")
        out = file.read()

        res.arquivo_txt = base64.b64encode(out)
        res.nome_txt = nome_arquivo_final

        file.close()

        os.remove(nome_arquivo_final)

    def retorno_converter_data_frame(self, tab):
        resultado = ''
        strcampos = ''
        array_campos = {}

        # CRIANDO ESTRUTURA DE DADOS DA CONVERTES
        for campos_convertes in tab.planilha_campo_ids.sorted('sequence'):
            array_campos.update({str(campos_convertes.mapeamento_id.name):
                                str})

        return array_campos


    def criando_tabela_sistema(self, tab):

        array_campos = []
        resultado = ''
        strcampos = ''

        # CRIANDO ESTRUTURA DE DADOS DA TABELA
        for campos in tab.planilha_campo_ids.sorted('sequence'):
            resultado = campos.mapeamento_id.name + ' ' + \
                        campos.mapeamento_id.tipo_campo
            if campos.mapeamento_id.tipo_campo != 'D':
                resultado += '(' + campos.mapeamento_id.tamanho_campo + ')'
            array_campos.append(resultado)

        strcampos = '; '.join([str(i) for i in array_campos])
        table = dbf.Table(tab.name, strcampos)
        table.open(mode=dbf.READ_WRITE)
        return table

    def ler_arquivo_excel(self, arquivo, tab):
        # ,converters={'COD_CIA': str, 'CPF_CNPJ': str}
        converters = self.retorno_converter_data_frame(tab)
        arquivo_lido = self._ler_arquivo(arquivo, tab.name, converters)
        arquivo_lido.iterrows()
        return arquivo_lido

    def criar_registro_operacional(self, res, arquivo):
        susep = SUSEP()
        for tab in res.quadro_id.planilha_ids:
            #criando tabela fisica dbf
            table = self.criando_tabela_sistema(tab)
            #lendo arquivo xls
            arquivo_lido = self.ler_arquivo_excel(arquivo, tab)

            #populando linhas para inserir no banco
            array_values_campos = []
            for index, col in arquivo_lido.iterrows():
                for campos in tab.planilha_campo_ids.sorted('sequence'):
                    name = campos.mapeamento_id.name
                    tipo_campo = campos.mapeamento_id.tipo_campo
                    valor = col[name]
                    valor = susep._mascaraCampos(valor,
                                                 campos.mapeamento_id.
                                                 formato1,
                                                 campos.mapeamento_id.
                                                 formato2,
                                                 campos.mapeamento_id.
                                                 trocar_valor)

                    valor = susep._convert_campos(tipo_campo, valor, name)

                    valor = '' if valor == 'nan' else valor

                    array_values_campos.append(valor)

                linha = tuple(array_values_campos)

                table.append(linha)

    def criar_quadros_txt(self, res, arquivo):
        susep = SUSEP()

        for tab in res.quadro_id.planilha_ids:
            # lendo arquivo xls
            arquivo_lido = self.ler_arquivo_excel(arquivo, tab)

            #lendo excel para gerar o TXT
            linha_completa=''
            for index, col in arquivo_lido.iterrows():
                resultado = ''
                for campos in tab.planilha_campo_ids.sorted('sequence'):
                    name = campos.mapeamento_id.name
                    tipo_campo = campos.mapeamento_id.tipo_campo
                    valor = col[name]
                    valor = susep._mascaraCampos(valor,
                                                 campos.mapeamento_id.
                                                 formato1,
                                                 campos.mapeamento_id.
                                                 formato2,
                                                 campos.mapeamento_id.
                                                 trocar_valor)

                    valor = susep._convert_campos(tipo_campo, valor, name)

                    valor = '' if valor == 'nan' else valor

                    resultado += valor

                linha_completa += resultado+'\n'

            nome_arquivo_final = res.quadro_id.nome_arquivo_exportacao + '.txt'
            arquivo = open(nome_arquivo_final, 'w+')
            arquivo.write(linha_completa)
            arquivo.close()

            file = open(nome_arquivo_final, "rb")
            out = file.read()

            res.arquivo_txt = base64.b64encode(out)

            res.nome_txt = nome_arquivo_final

            file.close()

            os.remove(nome_arquivo_final)




    @api.model
    def create(self, vals):
        arquivo = vals['arquivo']
        nome_arquivo = vals['nome_arquivo']

        if not arquivo and not nome_arquivo:
            raise UserError(u'É Necessário importar a planilha ')

        res = super(Susep, self).create(vals)

        if res.quadro_id.tipo_quadro == "1":
            self.criar_quadro_txt_susep(res, arquivo)
        elif res.quadro_id.tipo_quadro == "2":
            self.criar_quadros_txt(res, arquivo)
        else:
            self.criar_registro_operacional(res, arquivo)

        return res
